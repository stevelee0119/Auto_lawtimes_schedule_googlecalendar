import io
import json
import os
from datetime import datetime

import gspread
import openpyxl
from flask import Flask, render_template, request, send_file
from google.oauth2.service_account import Credentials

app = Flask(__name__)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]
SPREADSHEET_NAME = os.environ.get("SPREADSHEET_NAME", "인식표_정보수집")
ADMIN_SECRET = os.environ.get("ADMIN_SECRET", "")

HEADER = ["군번", "성명", "혈액형", "RH", "제출시간"]
BLOOD_TYPES = ["A", "B", "O", "AB"]


def get_sheet():
    creds_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if creds_json:
        creds_dict = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    else:
        creds = Credentials.from_service_account_file("credentials.json", scopes=SCOPES)

    client = gspread.authorize(creds)

    try:
        sheet = client.open(SPREADSHEET_NAME).sheet1
    except gspread.SpreadsheetNotFound:
        spreadsheet = client.create(SPREADSHEET_NAME)
        sheet = spreadsheet.sheet1
        sheet.append_row(HEADER)

    # Ensure header exists
    first_row = sheet.row_values(1) if sheet.row_count > 0 else []
    if first_row != HEADER:
        sheet.insert_row(HEADER, 1)

    return sheet


@app.route("/", methods=["GET"])
def form():
    return render_template("form.html", blood_types=BLOOD_TYPES)


@app.route("/submit", methods=["POST"])
def submit():
    gun_bun = request.form.get("gun_bun", "").strip()
    name = request.form.get("name", "").strip()
    blood_type = request.form.get("blood_type", "").strip()
    rh = request.form.get("rh", "").strip()

    errors = []
    if not gun_bun:
        errors.append("군번을 입력해주세요.")
    if not name:
        errors.append("성명을 입력해주세요.")
    if blood_type not in BLOOD_TYPES:
        errors.append("혈액형을 선택해주세요.")
    if rh not in ("RH+", "RH-"):
        errors.append("RH 여부를 선택해주세요.")

    if errors:
        return render_template("form.html", blood_types=BLOOD_TYPES, errors=errors,
                               prev={"gun_bun": gun_bun, "name": name,
                                     "blood_type": blood_type, "rh": rh})

    sheet = get_sheet()

    # Prevent duplicate submission by same 군번
    existing = sheet.col_values(1)  # 군번 column
    if gun_bun in existing:
        return render_template("form.html", blood_types=BLOOD_TYPES,
                               errors=["이미 제출된 군번입니다. 수정이 필요하면 담당자에게 문의하세요."],
                               prev={"gun_bun": gun_bun, "name": name,
                                     "blood_type": blood_type, "rh": rh})

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append_row([gun_bun, name, blood_type, rh, timestamp])

    return render_template("success.html", name=name)


@app.route("/download")
def download():
    secret = request.args.get("key", "")
    if ADMIN_SECRET and secret != ADMIN_SECRET:
        return "접근 권한이 없습니다.", 403

    sheet = get_sheet()
    data = sheet.get_all_values()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "인식표 정보"

    for i, row in enumerate(data):
        ws.append(row)
        if i == 0:
            # Bold header
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)

    # Auto-fit column widths (approximate)
    col_widths = [15, 10, 10, 8, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"인식표_정보_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
